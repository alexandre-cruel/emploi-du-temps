from __future__ import annotations

"""
Génération du fichier XLSX résultat.

Onglets :
  1. Planning groupes  — grille Jour × Spécialité/Groupe
  2. Planning par élève — une ligne par élève avec ses créneaux
  3. Effectifs          — récapitulatif doublettes, groupes, stats
"""

from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data import SLOTS, SPE_4_SLOTS, ParseResult
from solver import GroupResult, SolverResult

# Couleurs par spécialité
SPE_COLORS: dict[str, str] = {
    "Maths":  "D6E4F0",
    "SPC":    "FAD7A0",
    "SVT":    "A9DFBF",
    "SES":    "F9E79F",
    "HGGSP":  "D2B4DE",
    "HLP":    "FADBD8",
    "LCE":    "D5F5E3",
    "NSI":    "D6EAF8",
}
DEFAULT_COLOR = "EEEEEE"

HEADER_FILL = PatternFill("solid", fgColor="2E4057")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="7F8C8D")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF")


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _bold() -> Font:
    return Font(bold=True)


def _centered() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def generate_xlsx(result: SolverResult, parse_result: ParseResult) -> BytesIO:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore

    _sheet_planning_groupes(wb, result)
    _sheet_planning_eleves(wb, result, parse_result)
    _sheet_effectifs(wb, result, parse_result)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Onglet 1 : Planning groupes
# ---------------------------------------------------------------------------

def _sheet_planning_groupes(wb: openpyxl.Workbook, result: SolverResult) -> None:
    ws = wb.create_sheet("Planning groupes")

    groups = sorted(result.groups, key=lambda g: (g.specialite, g.groupe_id))

    # col_entries : (header_label, spe, slot_idx → cell_label)
    # Pour SPC/SVT : colonne par sous-groupe, avec label "cours" ou "TP"
    col_entries: list[tuple[str, str, dict[int, str]]] = []
    for g in groups:
        if g.specialite in SPE_4_SLOTS and g.subgroups and g.tp_pair:
            tp_a, tp_b = g.tp_pair
            for letter in ("A", "B"):
                slot_labels: dict[int, str] = {}
                for c in g.slots:
                    if c == tp_a:
                        slot_labels[c] = f"{g.label}{'A' if letter == 'A' else 'B'} TP"
                    elif c == tp_b:
                        # Seul le sous-groupe correspondant au slot c_b a cours ici
                        slot_labels[c] = f"{g.label}B TP" if letter == "B" else ""
                    else:
                        slot_labels[c] = f"{g.label} cours"
                col_entries.append((f"{g.label}{letter}", g.specialite, slot_labels))
        elif g.specialite in SPE_4_SLOTS and g.subgroups:
            # tp_pair pas détectée (rare) : fallback sans distinction
            for letter in ("A", "B"):
                col_entries.append((f"{g.label}{letter}", g.specialite, {c: g.label for c in g.slots}))
        else:
            col_entries.append((g.label, g.specialite, {c: g.label for c in g.slots}))

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18

    headers = ["Jour", "Créneau"] + [label for label, _, _ in col_entries]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = _centered()
        ws.column_dimensions[get_column_letter(col)].width = 18

    for row_idx, (slot_idx, day, start, end) in enumerate(SLOTS, start=2):
        ws.cell(row=row_idx, column=1, value=f"{day}").alignment = _centered()
        ws.cell(row=row_idx, column=2, value=f"{start}–{end}").alignment = _centered()

        for col_idx, (_, spe, slot_labels) in enumerate(col_entries, start=3):
            cell_label = slot_labels.get(slot_idx, "")
            if cell_label:
                color = SPE_COLORS.get(spe, DEFAULT_COLOR)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_label)
                cell.fill = _fill(color)
                cell.alignment = _centered()
                cell.font = _bold()

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Onglet 2 : Planning par élève
# ---------------------------------------------------------------------------

def _sheet_planning_eleves(
    wb: openpyxl.Workbook,
    result: SolverResult,
    parse_result: ParseResult,
) -> None:
    ws = wb.create_sheet("Planning par élève")

    slot_labels = [f"{day} {start}" for _, day, start, end in SLOTS]
    headers = ["Nom", "Prénom", "Doublette", "Options", "Sous-groupe"] + slot_labels + ["Nb créneaux"]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = _centered()

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 14
    for c in range(6, 6 + len(SLOTS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    student_groups = result.get_student_groups()
    student_slots = result.get_student_slots()

    # Map nom_prenom → sous-groupe letter (A ou B) pour SPC/SVT
    student_subgroup: dict[str, str] = {}
    for g in result.groups:
        if g.specialite in SPE_4_SLOTS and g.subgroups:
            for letter, members in g.subgroups.items():
                for st_obj in members:
                    student_subgroup[f"{st_obj.nom} {st_obj.prenom}"] = letter

    student_map = {f"{s.nom} {s.prenom}": s for s in parse_result.students}

    row_idx = 2
    for name in sorted(student_map.keys()):
        st_obj = student_map[name]
        grps = student_groups.get(name, [])
        slots = student_slots.get(name, [])

        ws.cell(row=row_idx, column=1, value=st_obj.nom)
        ws.cell(row=row_idx, column=2, value=st_obj.prenom)
        ws.cell(row=row_idx, column=3, value=" – ".join(st_obj.specialites))
        ws.cell(row=row_idx, column=4, value=", ".join(st_obj.options) if st_obj.options else "")
        ws.cell(row=row_idx, column=5, value=student_subgroup.get(name, ""))

        letter = student_subgroup.get(name, "")
        for slot_idx in slots:
            col = 6 + slot_idx
            spe_label = ""
            for g in grps:
                if slot_idx in g.slots:
                    if g.specialite in SPE_4_SLOTS and g.tp_pair:
                        tp_a, tp_b = g.tp_pair
                        if slot_idx == tp_a:
                            spe_label = f"{g.label}A TP" if letter == "A" else ""
                        elif slot_idx == tp_b:
                            spe_label = f"{g.label}B TP" if letter == "B" else ""
                        else:
                            spe_label = f"{g.label} cours"
                    else:
                        spe_label = g.label
                    break
            if not spe_label:
                continue
            cell = ws.cell(row=row_idx, column=col, value=spe_label)
            spe = spe_label.split()[0]
            cell.fill = _fill(SPE_COLORS.get(spe, DEFAULT_COLOR))
            cell.alignment = _centered()

        ws.cell(row=row_idx, column=6 + len(SLOTS), value=len(slots))
        row_idx += 1

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Onglet 3 : Effectifs
# ---------------------------------------------------------------------------

def _sheet_effectifs(
    wb: openpyxl.Workbook,
    result: SolverResult,
    parse_result: ParseResult,
) -> None:
    ws = wb.create_sheet("Effectifs")
    row = 1

    # ---- Résumé global ----
    ws.cell(row=row, column=1, value="RÉSUMÉ").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    ws.cell(row=row, column=1, value="Élèves total").font = _bold()
    ws.cell(row=row, column=2, value=result.stats.get("n_students", ""))
    row += 1
    ws.cell(row=row, column=1, value="Conflits de créneaux").font = _bold()
    ws.cell(row=row, column=2, value=result.stats.get("n_conflicts", ""))
    row += 1
    ws.cell(row=row, column=1, value="Élèves avec permanence").font = _bold()
    ws.cell(row=row, column=2, value=result.stats.get("n_permanences", ""))
    row += 1
    ws.cell(row=row, column=1, value="Statut solveur").font = _bold()
    ws.cell(row=row, column=2, value=result.status)
    row += 2

    # ---- Effectifs par groupe ----
    ws.cell(row=row, column=1, value="EFFECTIFS PAR GROUPE").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for h, col in [("Groupe", 1), ("Effectif", 2), ("Créneaux", 3)]:
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _fill("BDC3C7")
        cell.font = _bold()
    row += 1

    groups = sorted(result.groups, key=lambda g: (g.specialite, g.groupe_id))
    for g in groups:
        slot_str = ", ".join(
            f"{SLOTS[c][1]} {SLOTS[c][2]}" for c in g.slots
        )
        color = SPE_COLORS.get(g.specialite, DEFAULT_COLOR)
        for col, val in [(1, g.label), (2, g.effectif), (3, slot_str)]:
            cell = ws.cell(row=row, column=col, value=val)
            if col <= 2:
                cell.fill = _fill(color)
        row += 1
    row += 1

    # ---- Doublettes ----
    ws.cell(row=row, column=1, value="DOUBLETTES").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for h, col in [("Doublette", 1), ("Effectif", 2)]:
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _fill("BDC3C7")
        cell.font = _bold()
    row += 1
    for doublette, count in parse_result.doublette_counts.items():
        ws.cell(row=row, column=1, value=doublette)
        ws.cell(row=row, column=2, value=count)
        row += 1
    row += 1

    # ---- Effectifs par spécialité ----
    ws.cell(row=row, column=1, value="PAR SPÉCIALITÉ").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1
    for h, col in [("Spécialité", 1), ("Effectif", 2), ("Nb groupes", 3)]:
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _fill("BDC3C7")
        cell.font = _bold()
    row += 1
    for spe, count in parse_result.spe_counts.items():
        n_grp = len([g for g in result.groups if g.specialite == spe])
        color = SPE_COLORS.get(spe, DEFAULT_COLOR)
        for col, val in [(1, spe), (2, count), (3, n_grp)]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = _fill(color)
        row += 1

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40

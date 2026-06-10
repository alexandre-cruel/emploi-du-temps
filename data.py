from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import IO

import openpyxl

# ---------------------------------------------------------------------------
# Grille horaire fixe (Terminale et Première partagent la même grille)
# ---------------------------------------------------------------------------

SLOTS: list[tuple[int, str, str, str]] = [
    (0, "Lundi",    "13h55", "15h50"),
    (1, "Lundi",    "15h50", "17h50"),  # matex + DGEMC
    (2, "Mardi",    "8h10",  "10h00"),
    (3, "Mardi",    "10h15", "12h05"),
    (4, "Mercredi", "8h10",  "10h00"),  # matco
    (5, "Jeudi",    "8h10",  "10h00"),
    (6, "Jeudi",    "10h15", "12h05"),
    (7, "Vendredi", "8h10",  "10h00"),
    (8, "Vendredi", "10h15", "12h05"),
]

SLOT_LABELS = [f"J{i+1} – {day} {start}-{end}" for i, (_, day, start, end) in enumerate(SLOTS)]

# Slots réservés aux options (non assignables aux groupes de spé)
SLOT_MATEX = 1    # Lundi 15h50 — Maths expertes + DGEMC
SLOT_MATCO = 4    # Mercredi 8h10 — Maths complémentaires

N_SLOTS = len(SLOTS)

# Jours à double créneau (pour calcul des permanences)
DOUBLE_SLOT_PAIRS: list[tuple[int, int]] = [(2, 3), (5, 6), (7, 8)]

# Mapping normalisation des noms de spécialités (tels qu'ils apparaissent dans le fichier)
_SPE_ALIASES: dict[str, str] = {
    "M": "Maths", "MATHS": "Maths", "MATH": "Maths",
    "SPC": "SPC",
    "SVT": "SVT",
    "SES": "SES",
    "HG": "HGGSP", "HGGSP": "HGGSP",
    "HLP": "HLP",
    "LCE": "LCE",
    "NSI": "NSI",
}

# Spécialités nécessitant 4 créneaux (SPC et SVT en Terminale)
SPE_4_SLOTS: set[str] = {"SPC", "SVT"}

# Spécialités nécessitant 2 créneaux (mode Première — 3 spés × 2 créneaux)
# (dynamique : déterminé par niveau)

# Options valides par spécialité
# Maths expertes nécessite d'avoir Maths en spé
OPTION_MATEX_REQUIRES_MATHS = True


def normalize_spe(raw: str) -> str:
    """Normalise un nom de spécialité brut vers le nom canonique."""
    s = raw.strip().upper()
    return _SPE_ALIASES.get(s, raw.strip())


def parse_doublette(raw: str) -> list[str]:
    """
    Parse une chaîne comme '8) Maths - SPC' ou '3) HGGSP - SES'
    en ['Maths', 'SPC']. Gère aussi les triplettes pour la Première.
    """
    # Enlève le préfixe numérique "N) "
    s = re.sub(r"^\d+\)\s*", "", raw.strip())
    parts = [p.strip() for p in re.split(r"\s*[-–]\s*", s) if p.strip()]
    return [normalize_spe(p) for p in parts]


@dataclass
class Student:
    nom: str
    prenom: str
    classe_origine: str
    specialites: list[str]
    options: list[str]
    niveau: str  # "Terminale" | "Première"
    raw_doublette: str = ""

    @property
    def full_name(self) -> str:
        return f"{self.nom} {self.prenom}"

    @property
    def is_spc_svt(self) -> bool:
        return set(self.specialites) == {"SPC", "SVT"}


@dataclass
class ParseResult:
    students: list[Student]
    niveau: str
    doublette_counts: dict[str, int]
    spe_counts: dict[str, int]
    warnings: list[str]

    @property
    def all_specialites(self) -> list[str]:
        seen: set[str] = set()
        result: list[str] = []
        for s in self.students:
            for spe in s.specialites:
                if spe not in seen:
                    seen.add(spe)
                    result.append(spe)
        return sorted(result)


def parse_xlsx(source: str | IO[bytes] | BytesIO) -> ParseResult:
    """
    Parse un fichier XLSX de vœux élèves.
    Retourne un ParseResult avec tous les élèves et les statistiques.

    Colonnes attendues dans l'onglet 'Voeux élèves' :
      A: Nom, B: Prénom, C: Classe, D: Q1 (ignorée), E: Q2 (doublette), F: Q4 (option), G: Q5 (option2)
    """
    wb = openpyxl.load_workbook(source, data_only=True)

    if "Voeux élèves" not in wb.sheetnames:
        raise ValueError("Onglet 'Voeux élèves' introuvable dans le fichier.")

    ws = wb["Voeux élèves"]
    students: list[Student] = []
    warnings: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nom = row[0]
        prenom = row[1]
        classe = row[2]
        # row[3] = Q1 (situation), ignorée
        q2 = row[4]  # doublette
        q4 = row[5]  # option 1
        q5 = row[6]  # option 2

        # Ignore les lignes sans nom (annotations dans d'autres colonnes)
        if not nom or not isinstance(nom, str):
            continue

        # Ignore les élèves qui changent d'établissement
        q1 = row[3] or ""
        if "changement" in str(q1).lower():
            warnings.append(
                f"Ligne {row_idx} — {nom} {prenom or ''} : changement d'établissement, ignoré."
            )
            continue

        if not q2:
            warnings.append(f"Ligne {row_idx} — {nom} {prenom or ''} : doublette manquante, ignoré.")
            continue

        raw_doublette = str(q2).strip()

        # Filtre les élèves hors lycée (ex. "ST2S à Don Bosco")
        if "Don Bosco" in raw_doublette or "ST2S" in raw_doublette:
            warnings.append(f"Ligne {row_idx} — {nom} {prenom or ''} : orientation hors lycée ({raw_doublette}), ignoré.")
            continue

        try:
            specialites = parse_doublette(raw_doublette)
        except Exception as e:
            warnings.append(f"Ligne {row_idx} — {nom} {prenom or ''} : impossible de parser la doublette '{raw_doublette}': {e}")
            continue

        if len(specialites) < 2:
            warnings.append(f"Ligne {row_idx} — {nom} {prenom or ''} : doublette invalide '{raw_doublette}'.")
            continue

        # Options
        options: list[str] = []
        for opt_raw in [q4, q5]:
            if opt_raw and isinstance(opt_raw, str):
                opt = opt_raw.strip()
                if opt:
                    options.append(opt)

        # Détection niveau via classe_origine
        niveau = "Terminale"
        if classe and "Première" in str(classe):
            niveau = "Terminale"  # ce sont des Premières qui passent en Terminale l'an prochain

        # Validation : Maths expertes nécessite Maths en spé
        if OPTION_MATEX_REQUIRES_MATHS:
            if "Maths expertes" in options and "Maths" not in specialites:
                warnings.append(
                    f"{nom} {prenom or ''} : Maths expertes demandée sans Maths en spécialité — option retirée."
                )
                options = [o for o in options if o != "Maths expertes"]

        students.append(Student(
            nom=str(nom).strip(),
            prenom=str(prenom).strip() if prenom else "",
            classe_origine=str(classe).strip() if classe else "",
            specialites=specialites,
            options=options,
            niveau=niveau,
            raw_doublette=raw_doublette,
        ))

    if not students:
        raise ValueError("Aucun élève trouvé dans le fichier.")

    # Statistiques
    doublette_counts: dict[str, int] = {}
    spe_counts: dict[str, int] = {}
    for s in students:
        key = " – ".join(sorted(s.specialites))
        doublette_counts[key] = doublette_counts.get(key, 0) + 1
        for spe in s.specialites:
            spe_counts[spe] = spe_counts.get(spe, 0) + 1

    # Détection niveau global (Terminale si toutes les classes sont "Première Générale ...")
    # Les élèves du fichier sont des Premières qui passent en Terminale
    niveau_global = "Terminale"
    nb_spes = max((len(s.specialites) for s in students), default=2)
    if nb_spes == 3:
        niveau_global = "Première"

    return ParseResult(
        students=students,
        niveau=niveau_global,
        doublette_counts=dict(sorted(doublette_counts.items(), key=lambda x: -x[1])),
        spe_counts=dict(sorted(spe_counts.items(), key=lambda x: -x[1])),
        warnings=warnings,
    )
